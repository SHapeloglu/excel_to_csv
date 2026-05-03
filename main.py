"""
Excel → CSV Bölücü (Jinja2 Tabanlı) — Büyük Dosya Sürümü
----------------------------------------------------------
Beklenen Excel yapısı: 2 sütun → email | customData
• Boş satırları otomatik atlar
• openpyxl read_only ile RAM dostu chunk okuma
• concurrent.futures ile paralel CSV yazma
• Yerleşik ilerleme çubuğu (ETA dahil)
"""

import argparse
import math
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import openpyxl
from jinja2 import Environment, FileSystemLoader


# ── Ayarlar ──────────────────────────────────────────────────────────────────
ROWS_PER_FILE = 99
TEMPLATE_DIR  = Path(__file__).parent / "templates"
TEMPLATE_FILE = "csv_template.j2"
OUTPUT_DIR    = Path(__file__).parent / "output"
MAX_WORKERS   = min(8, (os.cpu_count() or 2) * 2)
# ─────────────────────────────────────────────────────────────────────────────


class ProgressBar:
    def __init__(self, total: int, label: str = ""):
        self.total = total
        self.done  = 0
        self.label = label
        self.start = time.time()
        self.width = 35

    def update(self, n: int = 1):
        self.done += n
        pct     = self.done / self.total if self.total else 1
        filled  = int(self.width * pct)
        bar     = "█" * filled + "░" * (self.width - filled)
        elapsed = time.time() - self.start
        eta     = (elapsed / pct - elapsed) if pct > 0 else 0
        sys.stdout.write(
            f"\r{self.label} [{bar}] "
            f"{self.done}/{self.total} ({pct*100:.1f}%)  ETA: {eta:.0f}s  "
        )
        sys.stdout.flush()
        if self.done >= self.total:
            sys.stdout.write("\n"); sys.stdout.flush()


def iter_excel_rows(excel_path: Path):
    """
    openpyxl read_only modunda satır satır okur.
    - İlk satırı başlık olarak alır
    - Tamamen boş satırları atlar
    - 2 sütun olduğunu doğrular
    Döndürür: (columns: list, row_generator)
    """
    wb  = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws  = wb.active
    rows = ws.iter_rows(values_only=True)

    # Başlık satırı
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    if len(header) != 2:
        wb.close()
        raise ValueError(
            f"Excel tam olarak 2 sütun içermelidir.\n"
            f"Bulunan sütunlar ({len(header)}): {header}"
        )

    print(f"  Sütunlar algılandı: {header[0]} | {header[1]}")

    def _gen():
        for row in rows:
            # Tamamen boş satırları atla
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            yield [str(c).strip() if c is not None else "" for c in row]
        wb.close()

    return header, _gen()


def count_nonempty_rows(excel_path: Path) -> int:
    """Boş olmayan veri satırlarını say (başlık hariç)."""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    count = 0
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if any(v is not None and str(v).strip() != "" for v in row):
            count += 1
    wb.close()
    return count


def render_and_write(task: dict) -> None:
    csv_text = task["template"].render(columns=task["columns"], rows=task["rows"])
    task["path"].write_text(csv_text, encoding="utf-8")


def flush_tasks(tasks: list, workers: int, progress: ProgressBar) -> None:
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futs = [pool.submit(render_and_write, t) for t in tasks]
        for f in as_completed(futs):
            f.result()
            progress.update()


def export(excel_path: str, output_dir: str = None,
           rows_per_file: int = ROWS_PER_FILE, workers: int = MAX_WORKERS):

    excel_path = Path(excel_path)
    output_dir = Path(output_dir) if output_dir else OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    env = Environment(
        loader=FileSystemLoader(str(TEMPLATE_DIR)),
        trim_blocks=True, lstrip_blocks=True, keep_trailing_newline=False,
    )
    template = env.get_template(TEMPLATE_FILE)

    print(f"\n📂 Kaynak   : {excel_path.name}")
    print(f"🔍 Satırlar sayılıyor...")
    total_rows = count_nonempty_rows(excel_path)
    file_count = math.ceil(total_rows / rows_per_file) if total_rows else 0
    base_name  = excel_path.stem

    columns, row_gen = iter_excel_rows(excel_path)

    print(f"📊 Toplam   : {total_rows:,} satır  |  Sütunlar: {columns}")
    print(f"📁 Çıktı    : {output_dir}")
    print(f"✂️  Parça    : {rows_per_file} satır/dosya  →  {file_count} dosya")
    print(f"⚡ Paralel  : {workers} thread\n")

    if total_rows == 0:
        print("⚠️  Veri satırı bulunamadı. İşlem sonlandırıldı.")
        return

    progress   = ProgressBar(total=file_count, label="Yazılıyor")
    tasks      = []
    file_index = 0
    batch      = []

    for row in row_gen:
        batch.append(row)
        if len(batch) == rows_per_file:
            file_index += 1
            tasks.append({
                "template": template, "columns": columns,
                "rows": batch,
                "path": output_dir / f"{base_name}_parca_{file_index:03d}.csv",
            })
            batch = []
            if len(tasks) >= workers * 4:
                flush_tasks(tasks, workers, progress)
                tasks = []

    if batch:
        file_index += 1
        tasks.append({
            "template": template, "columns": columns,
            "rows": batch,
            "path": output_dir / f"{base_name}_parca_{file_index:03d}.csv",
        })

    if tasks:
        flush_tasks(tasks, workers, progress)

    elapsed = time.time() - progress.start
    print(f"\n🎉 Tamamlandı!  {file_index} dosya  |  Süre: {elapsed:.1f}s  →  {output_dir}\n")


def main():
    parser = argparse.ArgumentParser(
        description="Excel (email | customData) → 99'ar satırlık CSV dosyaları"
    )
    parser.add_argument("excel",                  help="Kaynak Excel dosyası (ornek_veri.xlsx)")
    parser.add_argument("-o", "--output",          default=str(OUTPUT_DIR), help="Çıktı klasörü")
    parser.add_argument("-n", "--rows",   type=int, default=ROWS_PER_FILE,  help="Satır/dosya")
    parser.add_argument("-w", "--workers",type=int, default=MAX_WORKERS,    help="Paralel thread sayısı")
    args = parser.parse_args()
    export(args.excel, args.output, args.rows, args.workers)

if __name__ == "__main__":
    main()
